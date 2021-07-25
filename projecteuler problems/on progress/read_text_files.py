import time
from openpyxl import Workbook, load_workbook
f = open("Problem 745 output.txt", "r")

wb = Workbook()
ws = wb.active
ws.title = "Data"

i = 1
for line in f:
    parts = line.split("|")
    partpt = parts[2].split()
    partpt2 = partpt[2][:-4:]
    ws.cell(row = i, column = 2, value = partpt2)
    i+=1

wb.save('i.xlsx')